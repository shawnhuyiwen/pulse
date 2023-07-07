# Distributed under the Apache License, Version 2.0.
# See accompanying NOTICE file for details.

import sys
import json
import shutil
import logging
import numpy as np
from enum import Enum
from pathlib import Path
from dataclasses import dataclass, field
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from typing import Dict, Hashable, List, Tuple

from pulse.cdm.engine import SESegmentValidationTarget, SESegmentValidationTargetSegment
from pulse.cdm.scenario import SEScenario
from pulse.cdm.utils.file_utils import get_validation_dir
from pulse.cdm.io.engine import serialize_segment_validation_target_segment_to_file
from pulse.dataset.utils import generate_data_request


_pulse_logger = logging.getLogger('pulse')


def load_data(xls_file: Path) -> Tuple[str, str]:
    # Remove and recreate directory
    output_dir = Path("./validation/scenarios/")
    results_dir = Path("./test_results/scenarios/")
    xls_basename = "".join(xls_file.name.rsplit("".join(xls_file.suffixes), 1))
    xls_basename_out = xls_basename[:-10] if xls_basename.lower().endswith("validation") else xls_basename
    output_dir = output_dir / xls_basename_out
    results_dir = results_dir / xls_basename_out
    try:
        if output_dir.is_dir():
            shutil.rmtree(output_dir)
        output_dir.mkdir(parents=True, exist_ok=True)
    except OSError as e:
        _pulse_logger.error(f"Unable to clean directories")
        return

    _pulse_logger.info(f"Generating data from {xls_file}")

    # Iterate through each sheet in the file, generating a scenario for each
    workbook = load_workbook(filename=xls_file, data_only=True)
    for s in workbook.sheetnames:
        if s == "Notes":
            continue
        if not read_sheet(workbook[s], output_dir, results_dir):
            _pulse_logger.error(f"Unable to read {s} sheet")

    return output_dir, results_dir


# Read xlsx sheet and generate corresponding scenario file and validation target files
def read_sheet(sheet: Worksheet, output_dir: Path, results_dir: Path) -> bool:
    class Stage(Enum):
        IDScenario = 0
        InitialSegment = 1
        DataRequests = 2
        Segment = 3
        ValidationTargets = 4

    stage = Stage.IDScenario
    scenario = SEScenario()
    drs = []
    segments = []
    seg = None
    conditions = ""
    h2c = {}

    def _gen_header_to_col_dict(row: Tuple) -> Dict[str, int]:
        header_to_col = {}
        for col, h in enumerate(row):
            if h is None:
                continue
            header_to_col[h.strip().lower()] = col

        return header_to_col

    for row_num, r in enumerate(sheet.iter_rows(min_row=1, values_only=True)):
        # Scenario name and description
        if stage == Stage.IDScenario:
            # Header row
            if isinstance(r[0], str) and r[0].strip().lower() == "scenario name":
                h2c = _gen_header_to_col_dict(r)
                continue

            if "scenario name" in h2c:
                scenario.set_name(r[h2c["scenario name"]])
            if "patient file" in h2c and isinstance(r[h2c["patient file"]], str):
                scenario.get_patient_configuration().set_patient_file(r[h2c["patient file"]].strip())
            if "state file" in h2c and isinstance(r[h2c["state file"]], str):
                    scenario.set_engine_state(r[h2c["state file"]].strip())
            if "description" in h2c:
                scenario.set_description(r[h2c["description"]] if r[h2c["description"]] is not None else "")

            stage = Stage.InitialSegment
        # Segment 0
        elif stage == Stage.InitialSegment:
            # Header row
            if isinstance(r[0], str) and r[0].strip().lower() == "segment":
                h2c = _gen_header_to_col_dict(r)
                continue

            if not scenario.has_engine_state():
                if "conditions" in h2c and isinstance(r[h2c["conditions"]], str):
                    conditions = r[h2c["conditions"]]

            seg = SESegmentValidationTargetSegment()
            if "segment" in h2c:
                seg.set_segment_id(int(r[h2c["segment"]]))
            seg.set_notes(r[h2c["notes"]] if "notes" in h2c and isinstance(r[h2c["notes"]], str) else "")
            segments.append(seg)
            seg = None

            stage = Stage.DataRequests
        elif stage == Stage.DataRequests:
            # Header row
            if isinstance(r[0], str) and r[0].strip().lower() == "request type":
                h2c = _gen_header_to_col_dict(r)
                continue

            # Header row
            if isinstance(r[0], str) and r[0].strip().lower() == "segment":
                h2c = _gen_header_to_col_dict(r)
                stage = Stage.Segment
                continue

            if r[0] is not None:
                dr = generate_data_request(
                    request_type=r[h2c["request type"]] if "request type" in h2c and isinstance(r[h2c["request type"]], str) else "",
                    property_name=r[h2c["property name"]] if "property name" in h2c and isinstance(r[h2c["property name"]], str) else "",
                    unit_str=r[h2c["unit"]] if "unit" in h2c and isinstance(r[h2c["unit"]], str) else "",
                    precision=r[h2c["precision"]],
                )
                drs.append(dr)

            if "data request files" in h2c and isinstance(r[h2c["data request files"]], str):
                scenario.get_data_request_files().append(r[h2c["data request files"]].strip() + ".json")
        # Segment ID, note, and actions
        elif stage == Stage.Segment:
            # Header row
            if isinstance(r[0], str) and r[0].strip().lower() == "segment":
                h2c = _gen_header_to_col_dict(r)
                continue

            seg = SESegmentValidationTargetSegment()
            if "segment" in h2c:
                seg.set_segment_id(int(r[h2c["segment"]]))
            seg.set_notes(r[h2c["notes"]] if "notes" in h2c and isinstance(r[h2c["notes"]], str) else "")
            seg.set_actions(r[h2c["actions"]] if "actions" in h2c and isinstance(r[h2c["actions"]], str) else "")

            stage = Stage.ValidationTargets
        # Validation targets, comparison type/to, reference/notes
        elif stage == Stage.ValidationTargets:
            # Header row
            if isinstance(r[0], str) and r[0].strip().lower() == "request type":
                h2c = _gen_header_to_col_dict(r)
                continue

            # Header row, identified end of segment
            if isinstance(r[0], str) and r[0].strip().lower() == "segment":
                segments.append(seg)
                seg = None

                h2c = _gen_header_to_col_dict(r)
                stage = Stage.Segment
                continue

            if "request type" in h2c and r[h2c["request type"]] is not None:
                dr = generate_data_request(
                    request_type=r[h2c["request type"]] if "request type" in h2c and isinstance(r[h2c["request type"]], str) else "",
                    property_name=r[h2c["property name"]] if "property name" in h2c and isinstance(r[h2c["property name"]], str) else "",
                    unit_str=r[h2c["unit"]] if "unit" in h2c and isinstance(r[h2c["unit"]], str) else "",
                    precision=None,
                )
                val_tgt = SESegmentValidationTarget()
                val_tgt.set_header(dr.to_string())
                if isinstance(r[h2c["reference"]], str):
                    val_tgt.set_reference(r[h2c["reference"]])
                type_str = r[h2c["comparison type"]].strip().lower()
                comparison_str = r[h2c["comparison segment/value"]]
                if type_str == "equaltosegment":
                    val_tgt.set_equal_to_segment(int(comparison_str))
                elif type_str == "equaltovalue":
                    val_tgt.set_equal_to_value(float(comparison_str))
                elif type_str == "greaterthansegment":
                    val_tgt.set_greater_than_segment(int(comparison_str))
                elif type_str == "greaterthanvalue":
                    val_tgt.set_greater_than_value(float(comparison_str))
                elif type_str == "lessthansegment":
                    val_tgt.set_less_than_segment(int(comparison_str))
                elif type_str == "lessthanvalue":
                    val_tgt.set_less_than_value(float(comparison_str))
                elif type_str == "trendstosegment":
                    val_tgt.set_trends_to_segment(int(comparison_str))
                elif type_str == "trendstovalue":
                    val_tgt.set_trends_to_value(float(comparison_str))
                elif type_str == "range":
                    range_list = [float(val.strip()) for val in comparison_str[1:-1].split(",")]
                    val_tgt.set_range(range_list[0], range_list[1])
                elif type_str == "notvalidating":
                    pass
                elif type_str == "tbd":
                    _pulse_logger.warning("Found tbd target")
                else:
                    raise ValueError(f"Unable to identify comparison type: {type_str}")

                if "notes" in h2c:
                    val_tgt.set_notes(r[h2c["notes"]] if isinstance(r[h2c["notes"]], str) else "")

                seg.add_validation_target(val_tgt)
        else:
            raise ValueError(f"Unknown automated scenario validation stage: {stage}")

    if seg is not None:
        segments.append(seg)

    scenario.get_data_request_manager().set_data_requests(drs)
    full_results_dir = results_dir / scenario.get_name()
    scenario.get_data_request_manager().set_results_filename(str(full_results_dir / f"{scenario.get_name()}.csv"))

    # Write out scenario
    full_output_dir = output_dir / scenario.get_name()

    full_output_dir.mkdir(parents=True, exist_ok=True)
    filename = full_output_dir / (scenario.get_name() + ".json")
    _pulse_logger.info(f"Writing {filename}")
    with open(filename, "w") as f:
        f.write(write_scenario(scenario, segments, conditions, full_results_dir))

    # Write out validation target files
    for s in segments:
        if not s.has_validation_targets():
            continue
        filename = full_output_dir / f"Segment{s.get_segment_id()}ValidationTargets.json"
        _pulse_logger.info(f"Writing {filename}")
        serialize_segment_validation_target_segment_to_file(s, filename)

    return True


def write_scenario(scenario: SEScenario, segments: List[SESegmentValidationTargetSegment],
                   conditions: str, results_dir: Path) -> str:
    # Load actions into dict from concatenated JSON across segments
    all_actions_str = '{"AnyAction": ['
    for idx, s in enumerate(segments):
        all_actions_str += s.get_actions()

        # Add serialize requested action to end of every segment
        if idx == len(segments) - 1 and all_actions_str != '{"AnyAction": [':
            all_actions_str += ','
        all_actions_str += '{"SerializeRequested": {'
        segment_file_name = results_dir / f"Segment{s.get_segment_id()}.json"
        all_actions_str += f'"Filename": "{segment_file_name}"'
        all_actions_str += '}}'
        if idx != len(segments) - 1:
            all_actions_str += ','
    all_actions_str += ']}'
    all_actions = []
    if all_actions_str != '{"AnyAction": []}':
        all_actions = json.loads(all_actions_str)["AnyAction"]

    # Load conditions into dict
    all_conditions_str = '{"AnyCondition": ['
    all_conditions_str += conditions
    all_conditions_str += ']}'
    all_conditions = {}
    if conditions:
        all_conditions = json.loads(all_conditions_str)

    # Create patient configuration dict
    if scenario.has_patient_configuration():
        patient_config = scenario.get_patient_configuration()
        patient_config_dict = {}
        if patient_config.has_patient_file():
            patient_config_dict["PatientFile"] = patient_config.get_patient_file()
        elif patient_config.has_patient():
            raise ValueError("Patient to JSON not implemented")
        if all_conditions:
            patient_config_dict["Conditions"] = all_conditions
        if patient_config.get_data_root_dir() and patient_config.get_data_root_dir() != "./":
            patient_config_dict["DataRoot"] = patient_config.get_data_root_dir()

    # Create data request manager dict
    dr_mgr = scenario.get_data_request_manager()
    dr_mgr_dict = {}
    if dr_mgr.get_data_requests():
        dr_mgr_dict["DataRequest"] = []
        for dr in dr_mgr.get_data_requests():
            dr_dict = {}
            dfmt_dict = {}
            if dr.get_precision():
                dfmt_dict["Precision"] = dr.get_precision()
            if dr.has_notation():
                dfmt_dict["Type"] = dr.get_notation().name
            if dfmt_dict:
                dr_dict["DecimalFormat"] = dfmt_dict
            dr_dict["Category"] = dr.get_category().name
            if dr.has_action_name():
                dr_dict["ActionName"] = dr.get_action_name()
            if dr.has_compartment_name():
                dr_dict["CompartmentName"] = dr.get_compartment_name()
            if dr.has_substance_name():
                dr_dict["SubstanceName"] = dr.get_substance_name()
            if dr.has_property_name():
                dr_dict["PropertyName"] = dr.get_property_name()
            if dr.has_unit():
                dr_dict["Unit"] = dr.get_unit()

            dr_mgr_dict["DataRequest"].append(dr_dict)
    if dr_mgr.get_results_filename():
        dr_mgr_dict["ResultsFilename"] = dr_mgr.get_results_filename()

    # Compose full scenario dict
    scenario_dict = {}
    if scenario.has_name():
        scenario_dict["Name"] = scenario.get_name()
    if scenario.has_description():
        scenario_dict["Description"] = scenario.get_description()
    if scenario.has_patient_configuration():
        scenario_dict["PatientConfiguration"] = patient_config_dict
    elif scenario.has_engine_state():
        scenario_dict["EngineStateFile"] = scenario.get_engine_state()
    if scenario.get_data_request_files():
        scenario_dict["DataRequestFile"] = scenario.get_data_request_files()
    if dr_mgr_dict:
        scenario_dict["DataRequestManager"] = dr_mgr_dict
    if all_actions:
        scenario_dict["AnyAction"] = all_actions

    config_dict = {"TimeStep": {"ScalarTime": {"Value": 0.02, "Unit": "s"}}}

    out_dict = {
        "Configuration": config_dict,
        "Scenario": scenario_dict
    }

    return json.dumps(out_dict, indent=2)


if __name__ == "__main__":
    logging.basicConfig(level=logging.DEBUG, format='%(levelname)s: %(message)s')

    xls_file = None

    if len(sys.argv) < 2:
        _pulse_logger.error("Expected inputs : <xls validation file path>")
        sys.exit(1)

    xls_file = Path(sys.argv[1])
    if not xls_file.is_file():
        xls_file = Path(Path(get_validation_dir()) / xls_file)
        if not xls_file.is_file():
            _pulse_logger.error("Please provide a valid xls file")
            sys.exit(1)

    _ = load_data(xls_file)
