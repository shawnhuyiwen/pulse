# Distributed under the Apache License, Version 2.0.
# See accompanying NOTICE file for details.

import sys
import time
import shutil
import logging
import numbers
import numpy as np
from pathlib import Path
from dataclasses import dataclass
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils.cell import get_column_letter
from typing import Optional, Union
from pycel import ExcelCompiler

from pulse.cdm.engine import SETimeSeriesValidationTarget
from pulse.cdm.patient import SEPatient, eSex
from pulse.cdm.scalars import LengthUnit, MassUnit
from pulse.cdm.utils.file_utils import get_data_dir
from pulse.cdm.io.engine import serialize_time_series_validation_target_list_to_file
from pulse.cdm.io.patient import serialize_patient_from_file
from pulse.dataset.utils import generate_data_request

_pulse_logger = logging.getLogger('pulse')

def load_data(xls_file: Path):
    # Remove and recreate directory
    output_dir = Path("./validation/")
    xls_basename = "".join(xls_file.name.rsplit("".join(xls_file.suffixes), 1))
    xls_basename_out = xls_basename[:-4] if xls_basename.lower().endswith("data") else xls_basename
    output_dir = output_dir / xls_basename_out
    try:
        if output_dir.is_dir():
            shutil.rmtree(output_dir)
        output_dir.mkdir(parents=True, exist_ok=True)
    except OSError as e:
        _pulse_logger.error(f"Unable to clean directories")
        return

    if not xls_file.is_file():
        _pulse_logger.error(f"Could not find xls file {xls_file}")
        return

    _pulse_logger.info(f"Generating data from {xls_file}")

    # Copy file to preserve original state through patient updates
    timestr = time.strftime("%Y%m%d-%H%M%S")
    xls_edit = xls_file.parent / f"{xls_basename}_{timestr}.xlsx"

    # Iterate through patients
    patient_dir = Path("./patients/")
    try:
        if patient_dir.is_dir():
            for patient_file in patient_dir.glob("*.json"):
                update_patient(patient_file, xls_file, xls_edit)
                full_output_path = output_dir / patient_file.stem
                full_output_path.mkdir(parents=True, exist_ok=True)

                workbook = load_workbook(filename=xls_edit, data_only=False)
                evaluator = ExcelCompiler(filename=xls_edit)
                for system in workbook.sheetnames:
                    if system == "Patient":
                        continue
                    if not read_sheet(workbook[system], evaluator, full_output_path):
                        _pulse_logger.error(f"Unable to read {system} sheet")
    except:
        # Clean up temporary file no matter what
        if xls_edit.is_file():
            xls_edit.unlink()
        raise

    # Remove temp file
    if xls_edit.is_file():
        xls_edit.unlink()


def update_patient(patient_file: Path, xls_file: Path, new_file: Optional[Path]=None):
    p = SEPatient()
    serialize_patient_from_file(patient_file, p)

    workbook = load_workbook(filename=xls_file)
    sheet = workbook["Patient"]
    sheet["C2"] = "Male" if p.get_sex() == eSex.Male else "Female"
    if p.has_height():
        sheet["C3"] = p.get_height().get_value(units=LengthUnit.cm)
    if p.has_weight():
        sheet["C4"] = p.get_weight().get_value(units=MassUnit.kg)
    if p.has_right_lung_ratio():
        sheet["C9"] = p.get_right_lung_ratio().get_value()
    if p.has_body_fat_fraction():
        sheet["C12"] = p.get_body_fat_fraction().get_value()

    if not new_file:
        new_file = xls_file
    workbook.save(filename=new_file)


def read_sheet(sheet: Worksheet, evaluator: ExcelCompiler, output_dir: Path):
    system = sheet.title
    targets = {}

    # Get header to dataclass mapping
    ws_headers = [cell.value for cell in sheet[1]]
    try:
        VTB_HEADER = ws_headers.index('Output')
        VTB_UNITS = ws_headers.index('Units')
        VTB_ALGO = ws_headers.index('Algorithm')
        VTB_REF_CELL = ws_headers.index('Reference Values')
        VTB_TGT_FILE = ws_headers.index('ValidationTargetFile')
    except ValueError as e:
        _pulse_logger.error(f"Missing required header {str(e)[:str(e).find(' is not in list')]}")
        return False


    @dataclass
    class ValidationTargetBuilder():
        header: str
        units: str
        algorithm: str
        ref_cell: Union[str, numbers.Number]
        tgt_file: str

    for row_num, r in enumerate(sheet.iter_rows(min_row=2, values_only=True)):
        vtb = ValidationTargetBuilder(
            header = r[VTB_HEADER],
            units = r[VTB_UNITS],
            algorithm = r[VTB_ALGO],
            ref_cell=r[VTB_REF_CELL],
            tgt_file = r[VTB_TGT_FILE]
        )
        if not vtb.header:
            continue

        # Nothing to validate to
        if vtb.ref_cell is None:
            continue

        # TODO: Create targets for blank validation targets
        if not vtb.tgt_file or vtb.tgt_file == "ValidationTargetFile":
            continue

        if vtb.tgt_file not in targets:
            targets[vtb.tgt_file] = []
        vts = targets[vtb.tgt_file]

        prop_split = [s.strip() for s in vtb.header.split("-")]
        # TODO: Create other DR types
        dr = generate_data_request(
            request_type="LiquidCompartment",
            property_name=vtb.header.replace("*", ""),
            unit_str=vtb.units.strip(),
            precision=None)
        # TODO: References and notes
        tgt = SETimeSeriesValidationTarget()
        tgt.set_header(dr.to_string())

        ref_val = vtb.ref_cell

        algo = vtb.algorithm
        if algo == "Max":
            algo = "Maximum"
        elif algo == "Min":
            algo = "Minimum"
        algo = SETimeSeriesValidationTarget.eTargetType[algo]

        # Evaluate if needed
        if isinstance(ref_val, str) and ref_val.startswith("="):
            if ref_val.startswith("="):
                cell_loc = f"{system}!{get_column_letter(VTB_REF_CELL+1)}{row_num+2}"
                ref_val = evaluator.evaluate(cell_loc)

        # TODO: Support other comparison types
        if isinstance(ref_val, str):
            s_vals = ref_val.strip()
            s_vals = s_vals.replace('[', ' ')
            s_vals = s_vals.replace(']', ' ')
            vals = [float(s) for s in s_vals.split(',')]
            tgt.set_range(min(vals), max(vals), algo)
        elif isinstance(ref_val, numbers.Number):
            val = float(ref_val)
            tgt.set_equal_to(val, algo)
        else:
            _pulse_logger.warning(f"Unknown reference value type {ref_val}")
            tgt.set_range_min(np.nan)
            tgt.set_range_max(np.nan)

        vts.append(tgt)

    for target in targets.keys():
        filename = f"{output_dir}{system}{'-' if target else ''}{target}.json"
        _pulse_logger.info(f"Writing {filename}")
        serialize_time_series_validation_target_list_to_file(targets[target], filename)

    return True


if __name__ == "__main__":
    logging.basicConfig(level=logging.INFO, format='%(levelname)s: %(message)s')
    logging.getLogger("pycel").setLevel(logging.WARNING)

    xls_file = None

    if len(sys.argv) < 2:
        _pulse_logger.error("Expected inputs : <xls validation file path>")
        sys.exit(1)

    xls_file = Path(sys.argv[1])
    if not xls_file.is_file():
        xls_file = Path(get_data_dir()+sys.argv[1])
        if not xls_file.is_file():
            _pulse_logger.error("Please provide a valid xls file")
            sys.exit(1)

    load_data(xls_file)
