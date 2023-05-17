# Distributed under the Apache License, Version 2.0.
# See accompanying NOTICE file for details.

import logging
import numbers
import os
import shutil
import numpy as np
import time
from dataclasses import dataclass
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils.cell import get_column_letter
from typing import Optional, Union
from pycel import ExcelCompiler

from pulse.cdm.engine import SEDataRequestManager, SEValidationTarget, eValidationTargetType
from pulse.cdm.patient import SEPatient, eSex
from pulse.cdm.scalars import get_unit, LengthUnit, MassUnit
from pulse.cdm.utils.csv_utils import read_csv_into_df
from pulse.cdm.utils.file_utils import get_data_dir

from pulse.cdm.io.engine import serialize_data_request_manager_to_file
from pulse.cdm.io.patient import serialize_patient_from_file


_pulse_logger = logging.getLogger('pulse')


def load_data(xls_file: str):
    # Remove and recreate directory
    output_dir = "./validation/"
    xls_basename = os.path.splitext(os.path.basename(xls_file))[0]
    xls_basename_out = xls_basename[:-4] if xls_basename.lower().endswith("data") else xls_basename
    output_dir = os.path.join(output_dir, xls_basename_out + "/")
    try:
        if os.path.isdir(output_dir):
            shutil.rmtree(output_dir)
        os.makedirs(output_dir)
    except OSError as e:
        _pulse_logger.error(f"Unable to clean directories")
        return

    if not os.path.isfile(xls_file):
        _pulse_logger.error(f"Could not find xls file {xls_file}")
        return

    _pulse_logger.info(f"Generating data from {xls_file}")

    # Copy file to preserve original state through patient updates
    timestr = time.strftime("%Y%m%d-%H%M%S")
    xls_edit = os.path.join(os.path.dirname(xls_file), f"{xls_basename}_{timestr}.xlsx")

    # Iterate through patients
    patient_dir = "./patients/"
    try:
        if os.path.isdir(patient_dir):
            for patient_file in os.listdir(patient_dir):
                f = os.path.join(patient_dir, patient_file)
                if os.path.isfile(f) and patient_file.endswith(".json"):
                    update_patient(f, xls_file, xls_edit)
                    full_output_path = output_dir + patient_file[:-5] + "/"
                    os.makedirs(full_output_path)

                    workbook = load_workbook(filename=xls_edit, data_only=False)
                    evaluator = ExcelCompiler(filename=xls_edit)
                    for system in workbook.sheetnames:
                        if system == "Patient":
                            continue
                        if not read_sheet(workbook[system], evaluator, full_output_path):
                            _pulse_logger.error(f"Unable to read {system} sheet")
    except:
        # Clean up temporary file no matter what
        if os.path.isfile(xls_edit):
            os.remove(xls_edit)
        raise

    # Remove temp file
    if os.path.isfile(xls_edit):
        os.remove(xls_edit)


def update_patient(patient_file: str, xls_file: str, new_file: Optional[str]=None):
    p = SEPatient()
    serialize_patient_from_file(patient_file, p)

    workbook = load_workbook(filename=xls_file)
    sheet = workbook["Patient"]
    sheet["C2"] = "Male" if p.get_sex() == eSex.Male else "Female"
    if p.has_height():
        sheet["C3"] = p.get_height().get_value(units=LengthUnit.cm)
    if p.has_weight():
        sheet["C4"] = p.get_weight().get_value(units=MassUnit.kg)
    if p.has_right_lung_ratio():
        sheet["C9"] = p.get_right_lung_ratio().get_value()
    if p.has_body_fat_fraction():
        sheet["C12"] = p.get_body_fat_fraction().get_value()

    if not new_file:
        new_file = xls_file
    workbook.save(filename=new_file)


def read_sheet(sheet: Worksheet, evaluator: ExcelCompiler, output_dir: str):
    system = sheet.title
    targets = {}

    # Get header to dataclass mapping
    ws_headers = [cell.value for cell in sheet[1]]
    try:
        VTB_HEADER = ws_headers.index('Output')
        VTB_UNITS = ws_headers.index('Units')
        VTB_ALGO = ws_headers.index('Algorithm')
        VTB_REF_CELL = ws_headers.index('Reference Values')
        VTB_TGT_FILE = ws_headers.index('ValidationTargetFile')
    except ValueError as e:
        _pulse_logger.error(f"Missing required header {str(e)[:str(e).find(' is not in list')]}")
        return False


    @dataclass
    class ValidationTargetBuilder():
        header: str
        units: str
        algorithm: str
        ref_cell: Union[str, numbers.Number]
        tgt_file: str

    for row_num, r in enumerate(sheet.iter_rows(min_row=2, values_only=True)):
        vtb = ValidationTargetBuilder(
            header = r[VTB_HEADER],
            units = r[VTB_UNITS],
            algorithm = r[VTB_ALGO],
            ref_cell=r[VTB_REF_CELL],
            tgt_file = r[VTB_TGT_FILE]
        )
        if not vtb.header:
            continue

        # Nothing to validate to
        if vtb.ref_cell is None:
            continue

        # TODO: Create targets for blank validation targets
        if not vtb.tgt_file or vtb.tgt_file == "ValidationTargetFile":
            continue

        if vtb.tgt_file not in targets:
            targets[vtb.tgt_file] = SEDataRequestManager()
        dr_manager = targets[vtb.tgt_file]
        vts = dr_manager.get_validation_targets() if dr_manager.has_validation_targets() else []

        prop_split = [s.strip() for s in vtb.header.split("-")]
        unit = None
        try:
            unit = get_unit(vtb.units.strip())
        except:
            _pulse_logger.warning(f"Could not identify unit: {vtb.units}")
        # TODO: Create other DR types
        tgt = SEValidationTarget.create_liquid_compartment_validation_target(prop_split[0], prop_split[1], unit)
        tgt.set_type(eValidationTargetType[vtb.algorithm])


        ref_val = vtb.ref_cell

        # Evaluate if needed
        if isinstance(ref_val, str) and ref_val.startswith("="):
            if ref_val.startswith("="):
                cell_loc = f"{system}!{get_column_letter(VTB_REF_CELL+1)}{row_num+2}"
                ref_val = evaluator.evaluate(cell_loc)

        if isinstance(ref_val, str):
            s_vals = ref_val.strip()
            s_vals = s_vals.replace('[', ' ')
            s_vals = s_vals.replace(']', ' ')
            vals = [float(s) for s in s_vals.split(',')]
            tgt.set_range_min(min(vals))
            tgt.set_range_max(max(vals))
        elif isinstance(ref_val, numbers.Number):
            val = float(ref_val)
            tgt.set_range_min(val)
            tgt.set_range_max(val)
        else:
            _pulse_logger.warning(f"Unknown reference value type {ref_val}")
            tgt.set_range_min(np.nan)
            tgt.set_range_max(np.nan)

        vts.append(tgt)
        dr_manager.set_validation_targets(vts)

    for target in targets.keys():
        filename = f"{output_dir}{system}{'-' if target else ''}{target}.json"
        _pulse_logger.info(f"Writing {filename}")
        serialize_data_request_manager_to_file(targets[target], filename)

    return True


if __name__ == "__main__":
    logging.basicConfig(level=logging.INFO, format='%(levelname)s: %(message)s')
    logging.getLogger("pycel").setLevel(logging.WARNING)

    load_data(get_data_dir() + "/human/adult/validation/SystemValidationData.xlsx")
