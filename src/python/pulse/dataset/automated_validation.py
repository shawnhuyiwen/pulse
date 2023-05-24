# Distributed under the Apache License, Version 2.0.
# See accompanying NOTICE file for details.

import json
import logging
import numbers
import numpy as np
import os
import shutil
from dataclasses import dataclass, field
from enum import Enum
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from typing import Dict, Hashable, List, Optional, Tuple

from pulse.cdm.engine import SESegmentValidationTarget, eDataRequest_category, SEDataRequest
from pulse.cdm.scalars import get_unit
from pulse.cdm.scenario import SEScenario
from pulse.cdm.utils.file_utils import get_data_dir

from pulse.cdm.io.engine import serialize_segment_validation_target_list_to_file


_pulse_logger = logging.getLogger('pulse')


def load_data(xls_file: str) -> None:
    # Remove and recreate directory
    output_dir = "./validation/scenarios/"
    xls_basename = os.path.splitext(os.path.splitext(os.path.basename(xls_file))[0])[0]
    xls_basename_out = xls_basename[:-10] if xls_basename.lower().endswith("validation") else xls_basename
    output_dir = os.path.join(output_dir, xls_basename_out + "/")
    try:
        if os.path.isdir(output_dir):
            shutil.rmtree(output_dir)
        os.makedirs(output_dir)
    except OSError as e:
        _pulse_logger.error(f"Unable to clean directories")
        return

    _pulse_logger.info(f"Generating data from {xls_file}")

    # Iterate through each sheet in the file, generating a scenario for each
    workbook = load_workbook(filename=xls_file, data_only=True)
    for s in workbook.sheetnames:
        if s == "Notes":
            continue
        if not read_sheet(workbook[s], output_dir):
            _pulse_logger.error(f"Unable to read {s} sheet")


def generate_data_request(request_type: str, property_name: str, unit: str, precision: Optional[int]) -> SEDataRequest:
    request_type = request_type.strip()
    if request_type == "ActionCmpt" or request_type == "ActionSub":
        category = eDataRequest_category.Action
    else:
        category = eDataRequest_category[request_type.strip()]
    action = None
    compartment = None
    substance = None
    property = None
    unit = None if unit.strip().lower() == "unitless" else get_unit(unit)

    props = property_name.strip().split("-")

    if category == eDataRequest_category.Patient:
        property = property_name
    elif category == eDataRequest_category.Physiology:
        property = property_name
    elif category == eDataRequest_category.Environment:
        property = property_name
    elif category == eDataRequest_category.Action:
        if request_type == "ActionCmpt":
            action = props[0]
            compartment = props[1]
            property = props[2]
        elif request_type == "ActionSub":
            action = props[0]
            substance = props[1]
            property = props[2]
        else:
            action = props[0]
            property = props[1]
    elif category == eDataRequest_category.GasCompartment:
        compartment = props[0]
        if len(props) == 2:
            property = props[1]
        elif len(props) == 3:
            substance = props[1]
            property = props[2]
        else:
            raise ValueError(f"Invalid property name for GasCompartment Data Request: {property_name}")
    elif category == eDataRequest_category.LiquidCompartment:
        compartment = props[0]
        if len(props) == 2:
            property = props[1]
        elif len(props) == 3:
            substance = props[1]
            property = props[2]
        else:
            raise ValueError(f"Invalid property name for LiquidCompartment Data Request: {property_name}")
    elif category == eDataRequest_category.ThermalCompartment:
        compartment = props[0]
        property = props[1]
    elif category == eDataRequest_category.TissueCompartment:
        compartment = props[0]
        property = props[1]
    elif category == eDataRequest_category.Substance:
        substance = props[0]
        property = props[1]
    elif category == eDataRequest_category.AnesthesiaMachine:
        property = property_name
    elif category == eDataRequest_category.BagValveMask:
        property = property_name
    elif category == eDataRequest_category.ECG:
        property = property_name
    elif category == eDataRequest_category.ECMO:
        property = property_name
    elif category == eDataRequest_category.Inhaler:
        property = property_name
    elif category == eDataRequest_category.MechanicalVentilator:
        property = property_name
    else:
        raise ValueError(f"Unknown data request category: {request_type}")

    d_fmt = None
    if precision is not None:
        d_fmt = SEDecimalFormat()
        d_fmt.set_precision(precision)

    return SEDataRequest(category, action, compartment, substance, property, unit, d_fmt)


# Dataclass encapsulating information for a singular scenario "segment"
@dataclass
class ScenarioSegment():
    id: Hashable
    note: str = ""
    actions: str = ""
    val_tgts: List[SESegmentValidationTarget] = field(default_factory=list)


# Read xlsx sheet and generate corresponding scenario file and validation target files
def read_sheet(sheet: Worksheet, output_dir: str) -> bool:
    class Stage(Enum):
        IDScenario = 0
        InitialSegment = 1
        DataRequests = 2
        Segment = 3
        ValidationTargets = 4

    stage = Stage.IDScenario
    scenario = SEScenario()
    drs = []
    segments = []
    seg = None
    conditions = ""
    h2c = {}

    def _gen_header_to_col_dict(row: Tuple) -> Dict[str, int]:
        header_to_col = {}
        for col, h in enumerate(row):
            if h is None:
                continue
            header_to_col[h.strip().lower()] = col

        return header_to_col

    for row_num, r in enumerate(sheet.iter_rows(min_row=1, values_only=True)):
        # Scenario name and description
        if stage == Stage.IDScenario:
            # Header row
            if isinstance(r[0], str) and r[0].strip().lower() == "scenario name":
                h2c = _gen_header_to_col_dict(r)
                continue

            if "scenario name" in h2c:
                scenario.set_name(r[h2c["scenario name"]])
            if "patient file" in h2c and isinstance(r[h2c["patient file"]], str):
                scenario.get_patient_configuration().set_patient_file(r[h2c["patient file"]].strip())
            if "state file" in h2c and isinstance(r[h2c["state file"]], str):
                    scenario.set_engine_state(r[h2c["state file"]].strip())
            if "description" in h2c:
                scenario.set_description(r[h2c["description"]] if r[h2c["description"]] is not None else "")

            stage = Stage.InitialSegment
        # Segment 0
        elif stage == Stage.InitialSegment:
            # Header row
            if isinstance(r[0], str) and r[0].strip().lower() == "segment":
                h2c = _gen_header_to_col_dict(r)
                continue

            if not scenario.has_engine_state():
                if "conditions" in h2c and isinstance(r[h2c["conditions"]], str):
                    conditions = r[h2c["conditions"]]

            stage = Stage.DataRequests
        elif stage == Stage.DataRequests:
            # Header row
            if isinstance(r[0], str) and r[0].strip().lower() == "request type":
                h2c = _gen_header_to_col_dict(r)
                continue

            # Header row
            if isinstance(r[0], str) and r[0].strip().lower() == "segment":
                h2c = _gen_header_to_col_dict(r)
                stage = Stage.Segment
                continue

            if r[0] is not None:
                dr = generate_data_request(
                    request_type=r[h2c["request type"]] if "request type" in h2c and isinstance(r[h2c["request type"]], str) else "",
                    property_name=r[h2c["property name"]] if "property name" in h2c and isinstance(r[h2c["property name"]], str) else "",
                    unit=r[h2c["unit"]] if "unit" in h2c and isinstance(r[h2c["unit"]], str) else "",
                    precision=r[h2c["precision"]],
                )
                drs.append(dr)

            if "data request files" in h2c and isinstance(r[h2c["data request files"]], str):
                scenario.get_data_request_files().append(r[h2c["data request files"]].strip() + ".json")
        # Segment ID, note, and actions
        elif stage == Stage.Segment:
            # Header row
            if isinstance(r[0], str) and r[0].strip().lower() == "segment":
                h2c = _gen_header_to_col_dict(r)
                continue

            seg = ScenarioSegment(
                id = r[h2c["segment"]] if "segment" in h2c else "",
                note = r[h2c["notes"]] if "notes" in h2c and isinstance(r[h2c["notes"]], str) else "",
                actions = r[h2c["actions"]] if "actions" in h2c and isinstance(r[h2c["actions"]], str) else ""
            )

            stage = Stage.ValidationTargets
        # Validation targets, comparison type/to, reference/notes
        elif stage == Stage.ValidationTargets:
            # Header row
            if isinstance(r[0], str) and r[0].strip().lower() == "request type":
                h2c = _gen_header_to_col_dict(r)
                continue

            # Header row, identified end of segment
            if isinstance(r[0], str) and r[0].strip().lower() == "segment":
                segments.append(seg)
                seg = None

                h2c = _gen_header_to_col_dict(r)
                stage = Stage.Segment
                continue

            if "request type" in h2c and r[h2c["request type"]] is not None:
                dr = generate_data_request(
                    request_type=r[h2c["request type"]] if "request type" in h2c and isinstance(r[h2c["request type"]], str) else "",
                    property_name=r[h2c["property name"]] if "property name" in h2c and isinstance(r[h2c["property name"]], str) else "",
                    unit=r[h2c["unit"]] if "unit" in h2c and isinstance(r[h2c["unit"]], str) else "",
                    precision=None,
                )
                val_tgt = SESegmentValidationTarget()
                val_tgt.set_header(dr.to_string())
                if isinstance(r[h2c["reference"]], str):
                    val_tgt.set_reference(r[h2c["reference"]])
                tgt_str = r[h2c["target"]]
                tgt_str = tgt_str.strip().lower()
                comparison_segment = int(r[h2c["comparison segment"]])
                if tgt_str.startswith("equalto"):
                    val_tgt.set_equal_to(float(tgt_str.split(" ")[1]), comparison_segment)
                elif tgt_str.startswith("greaterthan"):
                    val_tgt.set_greater_than(float(tgt_str.split(" ")[1]), comparison_segment)
                elif tgt_str.startswith("lessthan"):
                    val_tgt.set_less_than(float(tgt_str.split(" ")[1]), comparison_segment)
                elif tgt_str == "increase":
                    val_tgt.set_increase(comparison_segment)
                elif tgt_str == "decrease":
                    val_tgt.set_decrease(comparison_segment)
                elif tgt_str.startswith("["):
                    range_list = [float(val.strip()) for val in tgt_str[1:-1].split(",")]
                    val_tgt.set_range(range_list[0], range_list[1], comparison_segment)
                elif tgt_str == "tbd":
                    _pulse_logger.warning("Found tbd target")
                else:
                    raise ValueError(f"Unable to identify comparison type: {r[h2c['target']]}")

                seg.val_tgts.append(val_tgt)
        else:
            raise ValueError(f"Unknown automated scenario validation stage: {stage}")

    if seg is not None:
        segments.append(seg)

    scenario.get_data_request_manager().set_data_requests(drs)

    # Write out scenario
    full_output_dir = os.path.join(output_dir, scenario.get_name())
    os.makedirs(full_output_dir)
    filename = os.path.join(full_output_dir, scenario.get_name() + ".json")
    _pulse_logger.info(f"Writing {filename}")
    with open(filename, "w") as f:
        f.write(write_scenario(scenario, segments, conditions))

    # Write out validation target files
    for s in segments:
        val_tgts = s.val_tgts
        filename = os.path.join(full_output_dir, f"Segment{s.id}ValidationTargets.json")
        _pulse_logger.info(f"Writing {filename}")
        serialize_segment_validation_target_list_to_file(val_tgts, filename)

    return True


def write_scenario(scenario: SEScenario, segments: List[ScenarioSegment], conditions: str) -> str:
    # Load actions into dict from concatenated JSON across segments
    all_actions_str = '{"AnyAction": ['
    for s in segments:
        all_actions_str += s.actions
    all_actions_str += ']}'
    all_actions = []
    if all_actions_str != '{"AnyAction": []}':
        all_actions = json.loads(all_actions_str)["AnyAction"]

    # Load conditions into dict
    all_conditions_str = '{"AnyCondition": ['
    all_conditions_str += conditions
    all_conditions_str += ']}'
    all_conditions = {}
    if conditions:
        all_conditions = json.loads(all_conditions_str)

    # Create patient configuration dict
    if scenario.has_patient_configuration():
        patient_config = scenario.get_patient_configuration()
        patient_config_dict = {}
        if patient_config.has_patient_file():
            patient_config_dict["PatientFile"] = patient_config.get_patient_file()
        elif patient_config.has_patient():
            raise ValueError("Patient to JSON not implemented")
        patient_config_dict["Conditions"] = all_conditions
        patient_config_dict["DataRoot"] = patient_config.get_data_root_dir()

    # Create data request manager dict
    dr_mgr = scenario.get_data_request_manager()
    dr_mgr_dict = {}
    if dr_mgr.get_data_requests():
        dr_mgr_dict["DataRequest"] = []
        for dr in dr_mgr.get_data_requests():
            dr_dict = {}
            dr_dict["DecimalFormat"] = {}
            if dr.get_precision():
                dr_dict["DecimalFormat"]["Precision"] = dr.get_precision()
            if dr.has_notation():
                dr_dict["DecimalFormat"]["Type"] = dr.get_notation().name
            dr_dict["Category"] = dr.get_category().name
            if dr.has_action_name():
                dr_dict["ActionName"] = dr.get_action_name()
            if dr.has_compartment_name():
                dr_dict["CompartmentName"] = dr.get_compartment_name()
            if dr.has_substance_name():
                dr_dict["SubstanceName"] = dr.get_substance_name()
            if dr.has_property_name():
                dr_dict["PropertyName"] = dr.get_property_name()
            if dr.has_unit():
                dr_dict["Unit"] = dr.get_unit()

            dr_mgr_dict["DataRequest"].append(dr_dict)
    if dr_mgr.get_results_filename():
        dr_mgr_dict["ResultFilename"] = dr_mgr.get_results_filename()

    # Compose full scenario dict
    scenario_dict = {}
    scenario_dict["Name"] = scenario.get_name()
    scenario_dict["Description"] = scenario.get_description()
    if scenario.has_patient_configuration():
        scenario_dict["PatientConfiguration"] = patient_config_dict
    elif scenario.has_engine_state():
        scenario_dict["EngineStateFile"] = scenario.get_engine_state()
    scenario_dict["DataRequestFile"] = scenario.get_data_request_files()
    scenario_dict["DataRequestManager"] = dr_mgr_dict
    scenario_dict["AnyAction"] = all_actions

    return json.dumps(scenario_dict, indent=2)


if __name__ == "__main__":
    logging.basicConfig(level=logging.DEBUG, format='%(levelname)s: %(message)s')

    load_data(get_data_dir() + "/human/adult/validation/Scenarios/AutomatedValidationTemplate.xlsx")
